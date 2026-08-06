"""XLSX and CSV extraction, with type coercion treated as a defect rather than a feature.

`docs/evaluation/DEAL_ROOM_EVAL.md` names the canonical failure directly: "Type coercion —
`000418` becoming `418` is the canonical failure." A customer id, a zip code, a GL account
and a part number are all strings that look like numbers, and a parser that helpfully
converts them destroys the join key the whole reconciliation depends on.

So nothing here ever calls `int()`, `float()` or `pd.read_csv`. Cell values are carried as
the source stored them, and a numeric interpretation is offered alongside the raw text
rather than replacing it.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path

import chardet
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from reef.extract.base import (
    ExtractionError,
    ExtractionResult,
    RawPage,
    RawSpan,
    SpanKind,
    TextSource,
)

#: A row wider than this is almost certainly a spreadsheet used as a canvas rather than a
#: table; extracting all of it produces spans that are mostly empty cells.
MAX_COLUMNS = 512


@dataclass(slots=True)
class CellValue:
    """A cell as stored, plus how it was stored. Both halves matter.

    `raw` is what the source contains and is what gets indexed and quoted. `numeric` is a
    convenience for deterministic calculation, and it is None whenever the source held
    text — including text that happens to look like a number, which is exactly the case
    that must not be silently converted.
    """

    address: str
    raw: str
    numeric: float | None
    is_text: bool


def _format_cell(cell: Cell, row_number: int, column_index: int) -> CellValue:
    # In read-only mode openpyxl yields `EmptyCell` for blanks, which carries no row or
    # column. Deriving the address from the iteration position works for both types and
    # keeps empty cells in their column, which matters because a blank is positional
    # information — "the FY2024 column is empty" is a finding.
    address = f"{get_column_letter(column_index)}{row_number}"
    value = getattr(cell, "value", None)

    if value is None:
        return CellValue(address=address, raw="", numeric=None, is_text=True)

    # `data_type == 's'` means the workbook stored a string. openpyxl already preserves
    # this correctly; the failure mode is downstream code that "cleans" it afterwards, so
    # `is_text` is recorded to make any later conversion an explicit decision.
    if getattr(cell, "data_type", None) == "s" or isinstance(value, str):
        return CellValue(address=address, raw=str(value), numeric=None, is_text=True)

    if isinstance(value, bool):
        return CellValue(
            address=address, raw="TRUE" if value else "FALSE", numeric=None, is_text=True
        )

    if isinstance(value, int | float):
        # `repr` on a float would render 22.4 as 22.4 but 0.1+0.2 as 0.30000000000000004.
        # Formatting through `str` keeps the human-facing value while `numeric` keeps full
        # precision for calculation.
        raw = str(int(value)) if isinstance(value, int) or value.is_integer() else str(value)
        return CellValue(address=address, raw=raw, numeric=float(value), is_text=False)

    # Dates and times arrive as datetime objects. ISO format is unambiguous; the
    # spreadsheet's display format is a presentation choice Reef does not inherit.
    return CellValue(address=address, raw=value.isoformat(), numeric=None, is_text=True)


def extract_xlsx(path: Path | None, data: bytes) -> ExtractionResult:
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            # Formulas are not evaluated. A cached formula result can disagree with the
            # formula, and reporting a stale cached value as fact is the kind of error
            # that survives review because it looks computed.
            data_only=True,
        )
    except Exception as exc:
        raise ExtractionError(f"workbook could not be opened: {exc}") from exc

    result = ExtractionResult(text_source=TextSource.NATIVE)
    cursor = 0
    sheet_names: list[str] = []

    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            sheet_names.append(sheet.title)
            # Each sheet is its own page, so a locator reads "sheet FY25 row 3" and the
            # eval's XLSX tolerance (sheet name exact, cell within the correct row) is
            # expressible as a single anchor.
            result.pages.append(RawPage(number=sheet_index, width=0.0, height=0.0))

            for row_number, row in enumerate(sheet.iter_rows(), start=1):
                cells = [
                    _format_cell(cell, row_number, column_index)
                    for column_index, cell in enumerate(row[:MAX_COLUMNS], start=1)
                ]
                if not any(c.raw for c in cells):
                    continue

                # Cells joined with a tab rather than a comma, so a value that itself
                # contains a comma does not create a phantom column boundary in the
                # indexed text.
                text = "\t".join(c.raw for c in cells).rstrip("\t")
                locator = f"sheet {sheet.title} row {row_number}"

                result.spans.append(
                    RawSpan(
                        page_number=sheet_index,
                        char_start=cursor,
                        char_end=cursor + len(text),
                        text=text,
                        locator=locator,
                        kind=SpanKind.TABLE_ROW,
                        breadcrumb=f"{sheet.title}",
                    )
                )
                cursor += len(text) + 1
    finally:
        workbook.close()

    result.extra = {
        "sheets": sheet_names,
        "page_count": len(result.pages),
    }
    if not result.spans:
        raise ExtractionError("workbook contains no populated cells")
    return result


def _decode(data: bytes) -> tuple[str, str]:
    """Decode text bytes, reporting which encoding was used.

    A CSV exported from Excel on Windows is frequently cp1252, and decoding it as UTF-8
    either raises or silently mangles every currency symbol and accented name in the file.
    """
    try:
        return data.decode("utf-8-sig"), "utf-8"
    except UnicodeDecodeError:
        pass
    detected = chardet.detect(data[:65536])
    encoding = detected.get("encoding") or "latin-1"
    try:
        return data.decode(encoding), encoding
    except (UnicodeDecodeError, LookupError):
        return data.decode("latin-1", errors="replace"), "latin-1"


def extract_csv(path: Path | None, data: bytes) -> ExtractionResult:
    text, encoding = _decode(data)

    sample = "\n".join(text.splitlines()[:20])
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    result = ExtractionResult(text_source=TextSource.NATIVE)
    result.pages.append(RawPage(number=1, width=0.0, height=0.0))
    cursor = 0
    header: list[str] = []

    # `csv.reader` returns every field as a string and never coerces. That is the entire
    # reason it is used here in preference to anything more convenient.
    reader = csv.reader(io.StringIO(text), dialect)
    row_count = 0
    for row_number, row in enumerate(reader, start=1):
        if not any(field.strip() for field in row):
            continue
        row_count += 1
        if row_number == 1:
            header = [field.strip() for field in row]

        line = "\t".join(row)
        result.spans.append(
            RawSpan(
                page_number=1,
                char_start=cursor,
                char_end=cursor + len(line),
                text=line,
                # Row index exact, at zero tolerance per the eval. 1-indexed with the
                # header as row 1, which is how a spreadsheet application numbers them
                # and therefore how a human reading the file will refer to it.
                locator=f"row {row_number}",
                kind=SpanKind.TABLE_ROW,
            )
        )
        cursor += len(line) + 1

    if not result.spans:
        raise ExtractionError("file contains no rows")

    result.extra = {
        "encoding": encoding,
        "delimiter": getattr(dialect, "delimiter", ","),
        "header": header,
        "row_count": row_count,
        "page_count": 1,
    }
    return result
